import openpyxl
import re
import json
import sys
import os
import time

HERE = os.path.dirname(os.path.abspath(__file__))
TIMETABLE_FILE = os.path.join(HERE, "timetable.xlsx")
FACULTY_FILE = os.path.join(HERE, "faculty.xlsx")
OUT_FILE = os.path.join(HERE, "data.js")

DAY_NAMES = {"MON", "TUES", "WED", "THUR", "FRI", "SAT"}

DEPT_NAME_TO_LETTER = {
    "CSE": "A", "CS": "A",
    "IT": "B",
    "BT": "C", "BIOTECH": "C", "BIOTECHNOLOGY": "C",
    "ECE": "G", "EC": "G",
}

MANUAL_ELECTIVES = {
    "C": {
        "CORE-AUDIT": [
            {"code": "20B13HS311", "name": "Indian Constitution & Traditional Knowledge",
             "credits": 3, "choice": 1, "compulsory": True},
        ],
        "CORE": [
            {"code": "15B11BT413", "name": "Bioprocess Engineering",
             "credits": 4, "choice": 1, "compulsory": True},
            {"code": "15B17BT472", "name": "Genetic Engineering Lab",
             "credits": 1, "choice": 1, "compulsory": True},
            {"code": "18B15BT311", "name": "Industrial Biotechnology Lab - I",
             "credits": 1, "choice": 1, "compulsory": True},
            {"code": "24B11BT311", "name": "Genetic Engineering",
             "credits": 4, "choice": 1, "compulsory": True},
            {"code": "24B17BT311", "name": "Summer Training-II (6 weeks)",
             "credits": 2, "choice": 1, "compulsory": True},
        ],
        "DE2": [
            {"code": "26B12BT314", "name": "Environmental Law: Science, Policy and Practice",
             "credits": 3, "choice": 1, "compulsory": True},
            {"code": "26B12BT312", "name": "Biocompliance and Safety Standards",
             "credits": 3, "choice": 2, "compulsory": False},
        ],
        "DE3": [
            {"code": "26B12BT315", "name": "Food Process Engineering",
             "credits": 3, "choice": 1, "compulsory": True},
            {"code": "26B12BT313", "name": "Molecular Oncology",
             "credits": 3, "choice": 2, "compulsory": False},
        ],
        "SE1": [
            {"code": "16B1NMA531", "name": "Discrete Mathematics",
             "credits": 3, "choice": 1, "compulsory": True},
            {"code": "16B1NMA533", "name": "Matrix Computations",
             "credits": 3, "choice": 2, "compulsory": False},
            {"code": "16B1NPH534", "name": "Bio-Materials Science",
             "credits": 3, "choice": 3, "compulsory": False},
            {"code": "17B1NMA531", "name": "Basic Numerical Methods",
             "credits": 3, "choice": 4, "compulsory": False},
        ],
    }
}

SHORT_CODE_NAMES = {
    "HS311": "Indian Constitution & Traditional Knowledge",
    "BT413": "Bioprocess Engineering",
    "BT472": "Genetic Engineering Lab",
    "BT314": "Environmental Law: Science, Policy and Practice",
    "BT312": "Biocompliance and Safety Standards",
    "BT315": "Food Process Engineering",
    "BT313": "Molecular Oncology",
    # Timetable cells sometimes drop the "B.." batch-year infix and use
    # <prefix><DEPT><suffix> instead of the full code (e.g. 18B15BT311 -> 18BT311).
    # These two short forms both point back to the same lab course.
    "18BT311": "Industrial Biotechnology Lab - I",
    "15BT311": "Industrial Biotechnology Lab - I",
    # Not present anywhere in the legend table itself; the sheet only labels
    # this slot "Minor Bioinformatics" next to the blank-code legend row.
    "PH412": "Minor Bioinformatics",
}

FULL_CODE_RE = re.compile(r'^(\d{2})B(\d{2})([A-Za-z]{2,4}\d{3,4})$')


def build_short_code_aliases(subjects):
    """Timetable cells often abbreviate a full code like 18B15BT311 by dropping
    the 'B..' infix, using either the leading 2 digits or the infix's 2 digits
    as a prefix (e.g. '18BT311' or '15BT311'). Derive both short aliases for
    every known full code so future subject-code variants resolve automatically.
    If two different full codes would collide on the same short alias, skip it
    rather than silently picking one (e.g. 24B11BT311 vs 24B17BT311 both reduce
    to '24BT311' -- ambiguous, so neither is aliased)."""
    candidates = {}
    for code, name in subjects.items():
        m = FULL_CODE_RE.match(code)
        if not m:
            continue
        prefix, infix, dept_suffix = m.groups()
        for alias in (f"{prefix}{dept_suffix}", f"{infix}{dept_suffix}"):
            candidates.setdefault(alias, set()).add(name)
    return {alias: next(iter(names)) for alias, names in candidates.items() if len(names) == 1}

BT311_BY_TYPE = {
    "P": "Industrial Biotechnology Lab - I",
    "L": "Genetic Engineering",
    "T": "Genetic Engineering",
}


def expand_numeric_batch_spec(batch_str, last_letter_hint=None):
    batch_str = batch_str.strip()
    if not batch_str:
        return []
    tokens = [t.strip() for t in batch_str.split(',') if t.strip()]
    result = []
    last_letter = last_letter_hint
    for tok in tokens:
        if '-' in tok:
            a, b = tok.split('-', 1)
            a, b = a.strip(), b.strip()
            ma = re.match(r'([A-Za-z]*)(\d+)', a)
            mb = re.match(r'([A-Za-z]*)(\d+)', b)
            if ma and mb:
                la = ma.group(1) or last_letter
                lb = mb.group(1) or la
                na, nb = int(ma.group(2)), int(mb.group(2))
                last_letter = lb
                for n in range(na, nb + 1):
                    result.append(f"{la}{n}")
            else:
                result.append(tok)
        else:
            m = re.match(r'([A-Za-z]*)(\d+)', tok)
            if m:
                letter = m.group(1) or last_letter
                last_letter = letter
                result.append(f"{letter}{m.group(2)}")
            else:
                result.append(tok)
    return result


def raw_cell_to_fields(text):
    text = str(text).strip()
    m = re.match(
        r'^(L|P|T)(minor)?\s*([A-Za-z0-9,\-\s]*)\(([^)]+)\)\s*/?\s*[-–]\s*/?\s*([^/]*?)\s*/\s*(.+)$',
        text
    )
    if not m:
        return None
    typ, minor, batch, course, room, fac = m.groups()
    return typ, bool(minor), batch.strip().rstrip(',').strip(), course.strip(), room.strip(), fac.strip()


def parse_elective_slot_code(text):
    text = str(text).strip()
    m = re.match(r'^(DE|SE)\s*(\d+)\s*/\s*(\d+)$', text)
    if not m:
        return None
    kind, num, choice = m.groups()
    return f"{kind}{num}", choice


def parse_free_text_elective(text):
    text = str(text).strip()
    if 'elective' not in text.lower():
        return None
    m = re.search(r'\(([^)]+)\)', text)
    if not m:
        return None
    dept_name = m.group(1).strip().upper()
    return DEPT_NAME_TO_LETTER.get(dept_name)


def get_day_blocks(rows):
    day_starts = [(i, r[0]) for i, r in enumerate(rows) if r[0] in DAY_NAMES]
    legend_row = next((i for i, r in enumerate(rows) if r[1] == 'SUBJECT CODE'), len(rows))
    blocks = []
    for idx, (start, name) in enumerate(day_starts):
        end = day_starts[idx + 1][0] if idx + 1 < len(day_starts) else legend_row
        blocks.append((name, start, end))
    return blocks, legend_row


def first_pass_collect_dept_batches(rows, blocks, times):
    dept_batches = {"A": set(), "B": set(), "C": set(), "G": set()}
    for day, start, end in blocks:
        for r in rows[start:end]:
            for col_idx, cell in enumerate(r[1:1 + len(times)]):
                if cell is None:
                    continue
                parsed = raw_cell_to_fields(cell)
                if not parsed:
                    continue
                _typ, minor, batch_spec, _course, _room, _fac = parsed
                if minor or not batch_spec:
                    continue
                if not re.search(r'\d', batch_spec):
                    continue  # letter-only spec, handled in second pass
                for b in expand_numeric_batch_spec(batch_spec):
                    m = re.match(r'^([A-Za-z])(\d+)$', b)
                    if m and m.group(1) in dept_batches:
                        dept_batches[m.group(1)].add(b)
    return {k: sorted(v, key=lambda x: int(x[1:])) for k, v in dept_batches.items()}


def parse_cell(text, day, time, dept_batches):
    text = str(text).strip()
    if not text or text.upper() == 'LUNCH':
        return None

    elective_slot = parse_elective_slot_code(text)
    if elective_slot:
        key, choice = elective_slot
        return {
            "day": day, "time": time, "type": key, "elective": True,
            "batches": [], "deptLetters": [], "course": "", "room": "",
            "faculty": "", "elective_choice": choice, "raw": text
        }

    free_elective_dept = parse_free_text_elective(text)
    if free_elective_dept:
        return {
            "day": day, "time": time, "type": "L", "elective": True,
            "batches": dept_batches.get(free_elective_dept, []),
            "deptLetters": [free_elective_dept], "course": "", "room": "",
            "faculty": "", "raw": text
        }

    parsed = raw_cell_to_fields(text)
    if not parsed:
        return {
            "day": day, "time": time, "type": "OTHER", "elective": False,
            "batches": [], "deptLetters": [], "course": "", "room": "",
            "faculty": "", "raw": text
        }

    typ, minor, batch_spec, course, room, fac = parsed

    if minor:
        return {
            "day": day, "time": time, "type": typ, "elective": False,
            "batches": ["minor"], "deptLetters": [], "course": course,
            "room": room, "faculty": fac, "raw": text
        }

    if batch_spec and re.search(r'\d', batch_spec):
        batches = expand_numeric_batch_spec(batch_spec)
        dept_letters = sorted({b[0] for b in batches if re.match(r'^[A-Za-z]\d+$', b)})
        return {
            "day": day, "time": time, "type": typ, "elective": False,
            "batches": batches, "deptLetters": dept_letters, "course": course,
            "room": room, "faculty": fac, "raw": text
        }

    if batch_spec and batch_spec.isalpha():
        letters = list(dict.fromkeys(batch_spec.upper()))
        batches = []
        for letter in letters:
            batches.extend(dept_batches.get(letter, []))
        return {
            "day": day, "time": time, "type": typ, "elective": True,
            "batches": batches, "deptLetters": letters, "course": course,
            "room": room, "faculty": fac, "raw": text
        }

    return {
        "day": day, "time": time, "type": typ, "elective": False,
        "batches": [], "deptLetters": [], "course": course,
        "room": room, "faculty": fac, "raw": text
    }


def parse_timetable(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    times = [t for t in rows[3][1:] if t]

    blocks, legend_row = get_day_blocks(rows)
    dept_batches = first_pass_collect_dept_batches(rows, blocks, times)

    entries = []
    for day, start, end in blocks:
        for r in rows[start:end]:
            for col_idx, cell in enumerate(r[1:1 + len(times)]):
                if cell is None:
                    continue
                e = parse_cell(cell, day, times[col_idx], dept_batches)
                if e:
                    entries.append(e)

    days = [d for d, _, _ in blocks]

    subjects = {}
    for r in rows[legend_row:legend_row + 40]:
        for i in (1, 4, 7):
            if i < len(r) and i + 1 < len(r) and r[i] and r[i + 1]:
                subjects[str(r[i]).strip()] = str(r[i + 1]).strip()

    current_key = {1: None, 4: None, 7: None}
    elective_legend = {}
    for r in rows[legend_row:legend_row + 40]:
        for ci in (1, 4, 7):
            code = r[ci] if ci < len(r) else None
            name = r[ci + 1] if ci + 1 < len(r) else None
            if code is None and name:
                label = str(name).strip()
                norm = re.sub(r'\s+', '', label).upper()
                if 'ELECTIVE-2' in norm:
                    current_key[ci] = 'DE2'
                elif 'ELECTIVE-3' in norm:
                    current_key[ci] = 'DE3'
                elif 'MINOR' in norm:
                    current_key[ci] = 'MINOR'
                else:
                    current_key[ci] = label
            elif code and name and current_key[ci]:
                elective_legend.setdefault(current_key[ci], [])
                elective_legend[current_key[ci]].append({
                    "code": str(code).strip(), "name": str(name).strip()
                })

    all_batches = sorted(
        {b for e in entries for b in e['batches'] if re.match(r'^[A-Za-z]\d+$', b)},
        key=lambda x: (x[0], int(x[1:]))
    )

    elective_keys = sorted({e['type'] for e in entries if re.match(r'^(DE|SE)\d+$', e['type'])})
    if any('minor' in e['batches'] for e in entries):
        elective_keys.append('minor')

    return entries, all_batches, elective_keys, times, days, subjects, elective_legend, dept_batches


def parse_faculty(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    faculty = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        for i in range(0, len(row), 2):
            code = row[i]
            name = row[i + 1] if i + 1 < len(row) else None
            if code and name and isinstance(code, str) and len(code) <= 6:
                faculty[code.strip()] = str(name).strip()
    return faculty


def build_once():
    if not os.path.exists(TIMETABLE_FILE):
        sys.exit(f"Missing {TIMETABLE_FILE} -- put your timetable Excel here as 'timetable.xlsx'")
    if not os.path.exists(FACULTY_FILE):
        sys.exit(f"Missing {FACULTY_FILE} -- put your faculty Excel here as 'faculty.xlsx'")

    entries, batches, electives, times, days, subjects, elective_legend, dept_batches = parse_timetable(TIMETABLE_FILE)
    faculty = parse_faculty(FACULTY_FILE)

    for slot in ("DE2", "DE3", "SE1"):
        if slot in MANUAL_ELECTIVES["C"]:
            elective_legend[slot] = MANUAL_ELECTIVES["C"][slot]

    # Manually curated names always win over the raw legend scrape (the sheet
    # itself has at least one duplicate-code typo, e.g. 24B11BT311 is listed
    # twice under two different names).
    for slot_list in MANUAL_ELECTIVES["C"].values():
        for item in slot_list:
            subjects[item["code"]] = item["name"]

    # Derive abbreviated aliases (e.g. 18B15BT311 -> 18BT311 / 15BT311) before
    # layering the hand-maintained short codes on top.
    subjects.update(build_short_code_aliases(subjects))
    subjects.update(SHORT_CODE_NAMES)

    # This planner is Biotech-only: keep entries relevant to C-batches, the
    # Biotech elective/minor slots, and drop everything belonging solely to
    # other departments (A/CSE, B/IT, G/ECE).
    def is_biotech_entry(e):
        if any(b.startswith("C") or b == "minor" for b in e["batches"]):
            return True
        if e["type"] in electives:  # DE2 / DE3 / SE1 placeholder rows
            return True
        return False

    entries = [e for e in entries if is_biotech_entry(e)]
    batches = [b for b in batches if b.startswith("C")]
    dept_batches = {"C": dept_batches.get("C", [])}

    used_courses = {e["course"] for e in entries if e["course"]}
    used_courses.add("BT311")
    subjects = {k: v for k, v in subjects.items() if k in used_courses}
    for item in MANUAL_ELECTIVES["C"].get("CORE-AUDIT", []) + MANUAL_ELECTIVES["C"].get("CORE", []):
        subjects.setdefault(item["code"], item["name"])

    used_faculty = set()
    for e in entries:
        for code in (e["faculty"] or "").split(","):
            code = code.strip()
            if code:
                used_faculty.add(code)
    faculty = {k: v for k, v in faculty.items() if k in used_faculty}

    payload = {
        "entries": entries,
        "faculty": faculty,
        "subjects": subjects,
        "batches": batches,
        "electives": electives,
        "elective_legend": elective_legend,
        "dept_batches": dept_batches,
        "bt311_by_type": BT311_BY_TYPE,
        "times": times,
        "days": days,
    }

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        f.write("const TT_DATA = " + json.dumps(payload, ensure_ascii=False, indent=1) + ";\n")

    print(f"data.js written: {len(entries)} entries, {len(batches)} batches, "
          f"{len(electives)} elective slots, {len(faculty)} faculty.")


def watch():
    print("Watching timetable.xlsx / faculty.xlsx for changes... (Ctrl+C to stop)")

    def mtimes():
        return {
            TIMETABLE_FILE: os.path.getmtime(TIMETABLE_FILE) if os.path.exists(TIMETABLE_FILE) else None,
            FACULTY_FILE: os.path.getmtime(FACULTY_FILE) if os.path.exists(FACULTY_FILE) else None,
        }

    last_seen = mtimes()
    build_once()
    try:
        while True:
            time.sleep(1.5)
            current = mtimes()
            if current != last_seen:
                last_seen = current
                try:
                    build_once()
                except SystemExit as e:
                    print(e)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    if "--watch" in sys.argv:
        watch()
    else:
        build_once()